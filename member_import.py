from es_connection import ElasticsearchConnect
from openpyxl import load_workbook

wb2 = load_workbook('member-20191020.xlsx', read_only=True)
print(wb2.sheetnames)
ws1 = wb2['议员']
con = ElasticsearchConnect()
es_client = con.connect()
for row in ws1.values:
	if row[2] == '' or row[2] is None:
		print("null值 跳过" + row[0])
		continue
	number_query = {
		"query": {
			"nested": {
				"path": "unique_item",
				"query": {
					"match_phrase": {
						"unique_item.value": row[2]
					}
				}
			}
		}
	}
	result = es_client.search(index="parsed_congress_members", _source=False, body=number_query)
	member_id = result['hits']['hits'][0]['_id']
	part = {"doc":{"representative_info" :{
		"id" : row[0],
		"served_histories" : row[6],
		"name" : row[1],
		"party" : row[7],
		"chamber" : row[8],
		"state" : row[10],
		"website" : row[11],
		"address" : row[12],
		"phone" : row[13],
		"district_address" : row[14],
		"district_phone" : row[15],
		"residence" : row[16],
		"marital_status" : row[17],
		"prev_occupation" : row[18],
		"prev_political_exp" : row[19],
		"education" : row[20],
		"birthdate" : row[21],
		"birthplace" : row[22],
		"religion" : row[23],
		"election_percentage" : row[24],
		"major_opponents" : row[25],
		"committeeList" : row[9]
	}}}
	res = es_client.update(index="parsed_congress_members", doc_type='parsed_congress_members_type', id=member_id, body=part)
	print(res)

# ws2 = wb2['助理团队']
# #i = 0
# for row in ws2.values:
#     try:
#         res1 = es_client.get(index="assistant_info", doc_type='assistant_info_type', id=row[0])
#         source = res1['_source']
#         new_member = {
#                             "name": row[3],
#                             "member_id": row[2],
#                             "job": row[4]
#                     }
#         source['service_member'].append(new_member)
#         res = es_client.index(index="assistant_info", doc_type='assistant_info_type', id=source['team_id'], body=source)
#         print('find data and ' + res['result'])
#     except:
#         doc = {
#             "team_id": row[0],
#             "name": row[1],
#             "service_member":[
#             {
#                     "name": row[3],
#                     "member_id": row[2],
#                     "job": row[4],
#                     }
#                     ],
#                     }
#         if row[5] != '' and row[5] is not None:
#             doc['twitter_account'] = row[5]
#         if row[7] != '' and row[7] is not None:
#             doc['linkedin_account'] = row[7]
#         if row[8] != '' and row[8] is not None:
#             doc['facebook_account'] = row[8]
#         res1 = es_client.index(index="assistant_info", doc_type='assistant_info_type', id=row[0], body=doc)
#         print('not find and ' + res1['result'])
        
