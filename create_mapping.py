from openpyxl import load_workbook
mapping = load_workbook('mapping_1105.xlsx', read_only=True)

ws = mapping['sheet1']
# three cache varible
member_name = ''
member_id = ''
member_sys_name = ''

# open news excel
senator_news = load_workbook('senator-news.xlsx')
num = 1
ws1 = senator_news['data']
for row in ws1.values:
    name = row[5]
    if name != member_name:
        for ws_row in ws.values:
            if ws_row[1] == name:
                member_name = ws_row[1]
                member_id = ws_row[2]
                member_sys_name = ws_row[3]
                break
    ws1.cell(row=num, column=7).value = member_id
    ws1.cell(row=num, column=8).value = member_sys_name
    num = num + 1
    print(num)


senator_news.save('senator-news.xlsx')
# ws.cell(row=1, column=6).value = "Jack"
# mapping.save('mapping_1105.xlsx')
#     print(row)
    # test_cell = row[5]
    # test_cell.value = 'jack'
    # if row[1] == member_name:
    #     print(row[1])
    #     print(row[2])
    #     print(row[3])
    #     print()


