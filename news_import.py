from es_connection import ElasticsearchConnect
from openpyxl import load_workbook
import time
import datetime

wb2 = load_workbook('senator-news.xlsx', read_only=True)
print(wb2.sheetnames)
ws = wb2['data']
con = ElasticsearchConnect()
es_client = con.connect()
for row in ws.values:
    s = row[3]
    time1 = int(time.mktime(datetime.datetime.strptime(s, "%Y-%m-%d %H:%M:%S").timetuple()) * 1000)
    news_body = {
		"title": row[0],
        "news_content": row[1],
        "media": row[2],
        "news_link": row[4],
        "member_name":row[7],
        "member_id":row[6],
        "release_time": time1,
	}

    res = es_client.index(index="news", doc_type='news_type', body=news_body)
    print(res)
print("news import complete!")