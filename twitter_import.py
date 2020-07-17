from es_connection import ElasticsearchConnect
from openpyxl import load_workbook

mapping = load_workbook('mapping.xlsx', read_only=True)
print(mapping.sheetnames)
con = ElasticsearchConnect()
es_client = con.connect()

ws = mapping['sheet1']
for row in ws.values:
    document_no = row[2]
    twitter_account = row[4]
    if twitter_account == '' or twitter_account is None:
        continue
    number_query = {
		"query": {
			"nested": {
				"path": "unique_item",
				"query": {
					"match_phrase": {
						"unique_item.value": document_no
					}
				}
			}
		}
	}
    result = es_client.search(index="parsed_congress_members", _source=False, body=number_query)
    member_id = result['hits']['hits'][0]['_id']
    part = {
        "doc" : {
            "twitter_account" : twitter_account
        }
    }
    res = es_client.update(index="parsed_congress_members", doc_type='parsed_congress_members_type', id=member_id, body=part)
    print(res)
